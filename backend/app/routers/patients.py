from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .. import phones
from ..db import calls, get_db
from ..security import current_user

router = APIRouter(prefix="/api/patients", tags=["patients"])

PATIENTS_COLLECTION = "patients"

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Column headings accepted in the upload. The first entry of each list is what the
# template ships with; the rest are aliases, so a sheet exported from a different
# system still imports without being reformatted by hand.
# visit_reason / visit_date were removed 2026-08-24: ChildLife confirmed they will not
# be supplying those two columns in the call list. The outbound agent no longer refers to
# the reason for the visit either — see _build_outbound_instructions in the agent repo.
COLUMN_ALIASES: dict[str, list[str]] = {
    "phone_number": ["phone number", "contact number", "phone", "contact", "mobile", "cell"],
    "patient_name": ["patient name", "name", "child name", "patient"],
    "mr_number": ["mr number", "mr#", "mr no", "mrn", "mr"],
    "er_name": ["er name", "er", "unit", "unit name", "hospital"],
    "patient_category": ["patient category", "category"],
    "disposition_category": ["disposition catg", "disposition category", "disposition"],
}

REQUIRED = ["phone_number"]


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def _map_columns(header_row: tuple) -> dict[int, str]:
    """Map each spreadsheet column index to a known field name."""
    lookup: dict[str, str] = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            lookup[alias] = field

    mapping: dict[int, str] = {}
    for index, cell in enumerate(header_row):
        field = lookup.get(_norm_header(cell))
        if field:
            mapping[index] = field
    return mapping


def _cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return re.sub(r"\s+", " ", str(value)).strip()


def patients():
    return get_db()[PATIENTS_COLLECTION]


@router.get("/template")
async def template(_: dict[str, Any] = Depends(current_user)) -> StreamingResponse:
    """Blank upload sheet with the expected headings and one example row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Patients"

    headings = [
        "Phone Number", "Patient Name", "MR Number", "ER name",
        "Patient Category", "Disposition Catg",
    ]
    ws.append(headings)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4A7FC0")
        cell.alignment = Alignment(horizontal="center")

    ws.append([
        "+923001234567", "Ali Khan", "MR-12345", "NIPA",
        "ER", "Sent Home",
    ])

    for i, h in enumerate(headings, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(len(h) + 6, 18)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="patient-upload-template.xlsx"'},
    )


@router.post("/upload")
async def upload(
    file: UploadFile = File(...),
    user: dict[str, Any] = Depends(current_user),
) -> dict[str, Any]:
    """Import patients from an .xlsx sheet.

    Rows are keyed on the normalized phone number, so re-uploading a corrected sheet
    updates the existing patients rather than creating duplicates.
    """
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    raw = await file.read()
    if len(raw) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File is larger than 10 MB")

    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read that file as a spreadsheet") from None

    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    try:
        header = next(rows)
    except StopIteration:
        raise HTTPException(status_code=400, detail="The sheet is empty") from None

    mapping = _map_columns(header)
    missing = [f for f in REQUIRED if f not in mapping.values()]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=(
                "The sheet needs a 'Phone Number' column. Found: "
                + (", ".join(str(h) for h in header if h) or "no headings")
            ),
        )

    batch_id = f"{datetime.now():%Y%m%d-%H%M%S}"
    collection = patients()
    await collection.create_index("phone_key", unique=True)
    await collection.create_index("call_status")

    created = updated = skipped = 0
    problems: list[dict[str, Any]] = []

    for line_no, row in enumerate(rows, start=2):
        record = {field: _cell(row[i]) for i, field in mapping.items() if i < len(row)}
        number = record.get("phone_number", "")

        if not any(record.values()):
            continue  # blank row

        if not phones.is_plausible(number):
            skipped += 1
            if len(problems) < 25:
                problems.append({
                    "row": line_no,
                    "value": number,
                    "reason": "Not a valid Pakistani mobile number",
                })
            continue

        key = phones.normalize(number)
        doc = {
            "phone_key": key,
            "phone_e164": phones.to_e164(number),
            "phone_variants": phones.variants(number),
            "phone_raw": number,
            "patient_name": record.get("patient_name", ""),
            "mr_number": record.get("mr_number", ""),
            "er_name": record.get("er_name", ""),
            "patient_category": record.get("patient_category", ""),
            "disposition_category": record.get("disposition_category", ""),
            "batch_id": batch_id,
            "uploaded_at": datetime.now(),
            "uploaded_by": user.get("username", ""),
            # Uploading someone is an explicit request to call them, so un-archive on
            # the way in. Without this, re-uploading a sheet whose numbers were
            # previously archived silently imports them straight back into the hidden
            # state and the queue stays empty for no visible reason.
            "archived": False,
        }

        result = await collection.update_one(
            {"phone_key": key},
            {
                "$set": doc,
                # Only set on insert, so re-uploading never resets a patient who has
                # already been called back to the top of the queue.
                "$setOnInsert": {"call_status": "pending", "attempts": 0, "last_called_at": None},
            },
            upsert=True,
        )
        if result.upserted_id:
            created += 1
        else:
            updated += 1

    wb.close()

    return {
        "batch_id": batch_id,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "problems": problems,
        "total_patients": await collection.count_documents({"archived": {"$ne": True}}),
    }


async def _attach_call_activity(items: list[dict[str, Any]]) -> None:
    """Fill in each patient's call history from conversation-logs.

    Done as one query for the whole page rather than per patient.
    """
    if not items:
        return

    all_variants: list[str] = []
    for p in items:
        all_variants.extend(p.get("phone_variants") or [])
    if not all_variants:
        return

    fields = [
        "patient_context.contact_number",
        "caller_number",
        "feedback_output.Phone Number",
    ]
    # Archived calls must be excluded here as well. This was the one query in the
    # backend that read conversation-logs without the filter, so a patient's "latest
    # call" could be an archived record — and since pre-2026-08-24 records stored
    # Pakistan wall-clock as if it were UTC, those sort ahead of calls that genuinely
    # happened later. A patient who had just gone unanswered showed as completed.
    query = {
        "archived": {"$ne": True},
        "$or": [{f: {"$in": all_variants}} for f in fields],
    }

    by_key: dict[str, list[dict[str, Any]]] = {}
    async for doc in calls().find(
        query,
        {"logs": 0, "metrics": 0},
    ).sort("timestamp", -1):
        pc = doc.get("patient_context") or {}
        fo = doc.get("feedback_output") or {}
        num = pc.get("contact_number") or doc.get("caller_number") or fo.get("Phone Number") or ""
        key = phones.normalize(num)
        if key:
            by_key.setdefault(key, []).append(doc)

    for p in items:
        history = by_key.get(p["phone_key"], [])
        p["call_count"] = len(history)
        latest = history[0] if history else None
        if latest:
            ts = latest.get("timestamp")
            p["last_call"] = {
                "at": ts.isoformat() if isinstance(ts, datetime) else None,
                "status": latest.get("status", ""),
                "session_id": latest.get("session_id", ""),
                "satisfied": latest.get("satisfied"),
            }
            # Reaching a real outcome is what closes a patient out; a silent or
            # unanswered call leaves them due for another attempt.
            p["derived_status"] = (
                "completed" if latest.get("status") in ("answered", "satisfied") else "attempted"
            )
        else:
            p["last_call"] = None
            p["derived_status"] = "pending"


def _clean(doc: dict[str, Any]) -> dict[str, Any]:
    doc["id"] = str(doc.pop("_id"))
    for k in ("uploaded_at", "last_called_at", "next_retry_at", "claimed_at"):
        if isinstance(doc.get(k), datetime):
            doc[k] = doc[k].isoformat()
    # Written by dialer.py; absent on rows uploaded before retries existed.
    doc.setdefault("attempts", 0)
    doc.setdefault("call_status", "pending")
    doc.setdefault("last_outcome", "")
    doc.setdefault("next_retry_at", None)
    return doc


@router.get("")
async def list_patients(
    search: str | None = None,
    status: str | None = Query(None, description="pending | attempted | completed"),
    er: str | None = None,
    batch_id: str | None = None,
    page: int = Query(1, ge=1),
    limit: int = Query(25, ge=1, le=200),
    _: dict[str, Any] = Depends(current_user),
) -> dict[str, Any]:
    query: dict[str, Any] = {"archived": {"$ne": True}}
    if er:
        query["er_name"] = er
    if batch_id:
        query["batch_id"] = batch_id
    if search:
        rx = re.compile(re.escape(search.strip()), re.IGNORECASE)
        query["$or"] = [
            {"patient_name": rx}, {"mr_number": rx},
            {"phone_e164": rx}, {"phone_raw": rx}, {"phone_key": rx},
        ]

    collection = patients()
    total = await collection.count_documents(query)
    cursor = (
        collection.find(query)
        .sort([("uploaded_at", -1), ("patient_name", 1)])
        .skip((page - 1) * limit)
        .limit(limit)
    )
    items = [_clean(d) async for d in cursor]
    await _attach_call_activity(items)

    # Status is derived from call history, so it is filtered after that lookup.
    if status:
        items = [p for p in items if p.get("derived_status") == status]

    return {
        "items": items,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": max(1, (total + limit - 1) // limit),
    }


@router.get("/queue")
async def queue(
    er: str | None = None,
    limit: int = Query(100, ge=1, le=500),
    include_attempted: bool = True,
    _: dict[str, Any] = Depends(current_user),
) -> dict[str, Any]:
    """Patients still due a call, oldest upload first."""
    query: dict[str, Any] = {"archived": {"$ne": True}}
    if er:
        query["er_name"] = er

    collection = patients()
    # Pull a generous slice, then filter on derived status — call history is what
    # decides who is still due, and that lives in another collection.
    items = [_clean(d) async for d in collection.find(query).sort("uploaded_at", 1).limit(limit * 4)]
    await _attach_call_activity(items)

    wanted = {"pending"} | ({"attempted"} if include_attempted else set())
    due = [p for p in items if p.get("derived_status") in wanted][:limit]

    counts = {"pending": 0, "attempted": 0, "completed": 0}
    for p in items:
        counts[p.get("derived_status", "pending")] = counts.get(p.get("derived_status", "pending"), 0) + 1

    # How the queue is distributed across retry attempts, so the dashboard can answer
    # "kitne retry pe hain" without the caller doing the arithmetic themselves.
    by_attempt: dict[str, int] = {}
    retries_waiting = 0
    now = datetime.now()
    for p in items:
        attempts = int(p.get("attempts") or 0)
        by_attempt[str(attempts)] = by_attempt.get(str(attempts), 0) + 1
        nxt = p.get("next_retry_at")
        if p.get("call_status") == "attempted" and nxt:
            when = datetime.fromisoformat(nxt) if isinstance(nxt, str) else nxt
            if when and when > now:
                retries_waiting += 1

    return {
        "items": due,
        "counts": counts,
        "by_attempt": by_attempt,
        "retries_waiting": retries_waiting,
        "total_due": len(due),
        "scanned": len(items),
    }


@router.get("/batches")
async def batches(_: dict[str, Any] = Depends(current_user)) -> list[dict[str, Any]]:
    """Upload history, newest first."""
    pipeline = [
        {"$match": {"archived": {"$ne": True}}},
        {"$group": {
            "_id": "$batch_id",
            "count": {"$sum": 1},
            "uploaded_at": {"$max": "$uploaded_at"},
            "uploaded_by": {"$first": "$uploaded_by"},
        }},
        {"$sort": {"uploaded_at": -1}},
        {"$limit": 30},
    ]
    out = []
    async for row in patients().aggregate(pipeline):
        ts = row.get("uploaded_at")
        out.append({
            "batch_id": row["_id"],
            "count": row["count"],
            "uploaded_at": ts.isoformat() if isinstance(ts, datetime) else None,
            "uploaded_by": row.get("uploaded_by", ""),
        })
    return out
