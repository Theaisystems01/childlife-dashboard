from __future__ import annotations

import io
from datetime import datetime, timedelta, timezone
from typing import Any

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..db import calls
from ..reporting import REPORT_COLUMNS, to_report_row
from ..security import current_user
from .calls import build_filter

router = APIRouter(prefix="/api/export", tags=["export"])

# Everything the foundation reads is Pakistan time. Mongo stores UTC, and the agent
# writes tz-aware UTC, so the conversion happens here rather than being left to whatever
# timezone the reader's machine happens to be in.
PKT = timezone(timedelta(hours=5))

# Legible in the cell, and still a real date underneath so Excel can sort and filter it.
DATE_FORMAT = "dd mmm yyyy  hh:mm AM/PM"


def _to_pkt_excel(value: Any) -> Any:
    """Turn a stored timestamp into a naive PKT datetime openpyxl will accept.

    openpyxl refuses timezone-aware datetimes, so the offset is applied and then
    dropped. Anything unparseable is passed through untouched rather than blanked —
    a slightly odd cell beats a silently empty one.
    """
    if isinstance(value, str) and value:
        try:
            value = datetime.fromisoformat(value)
        except ValueError:
            return value
    if not isinstance(value, datetime):
        return value
    if value.tzinfo is None:
        # Written before the agent used tz-aware timestamps; those were UTC.
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(PKT).replace(tzinfo=None)

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


@router.get("/xlsx")
async def export_xlsx(
    search: str | None = None,
    status: str | None = None,
    category: str | None = None,
    er: str | None = None,
    direction: str | None = None,
    days: int | None = None,
    complaints_only: bool = False,
    _: dict[str, Any] = Depends(current_user),
) -> StreamingResponse:
    """Export the current view in the foundation's exact report column order."""
    query = build_filter(search, status, category, er, direction, days)
    if complaints_only:
        query["feedback_summary.is_valid_feedback"] = True

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Feedback Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1C5CAB")
    sheet.append(REPORT_COLUMNS)
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    time_column = REPORT_COLUMNS.index("Received Time") + 1

    widths = [len(column) for column in REPORT_COLUMNS]
    row_number = 1
    async for doc in calls().find(query, {"logs": 0}).sort("timestamp", -1):
        row = to_report_row(doc)
        values = [row.get(column, "") for column in REPORT_COLUMNS]
        values[time_column - 1] = _to_pkt_excel(values[time_column - 1])
        sheet.append(values)
        row_number += 1

        cell = sheet.cell(row=row_number, column=time_column)
        if isinstance(cell.value, datetime):
            cell.number_format = DATE_FORMAT

        for index, value in enumerate(values):
            shown = value.strftime("%d %b %Y  %I:%M %p") if isinstance(value, datetime) else str(value)
            widths[index] = max(widths[index], min(len(shown), 60))

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width + 3

    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"childlife-feedback-{datetime.now(PKT):%Y%m%d-%H%M}.xlsx"
    return StreamingResponse(
        buffer,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
